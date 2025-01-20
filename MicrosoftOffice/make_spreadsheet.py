#!/usr/bin/env python3
"""Demonstrate some capabilities of openpyxl."""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart


def demonstrate():
    """Demonstrate openpyxl features."""
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo Sheet"

    # Adding data
    ws.append(["ID", "Name", "Score", "Grade"])
    students = [
        (1, "Alice", 85),
        (2, "Bob", 78),
        (3, "Charlie", 92),
        (4, "Diana", 88),
        (5, "Evan", 76),
    ]
    for student in students:
        ws.append(student)

    # Adding formulas
    ws["D2"] = (
        '=IF(C2>=90,"A",IF(C2>=80,"B",IF(C2>=70,"C","F")))'  # Grade based on score
    )
    for row in range(3, len(students) + 2):
        ws[f"D{row}"] = f'=IF(C{row}>=90,"A",IF(C{row}>=80,"B",IF(C{row}>=70,"C","F")))'

    # Formatting headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Adjusting column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Adding borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=2, max_row=len(students) + 1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # Adding a bar chart
    chart = BarChart()
    chart.title = "Student Scores"
    chart.x_axis.title = "Students"
    chart.y_axis.title = "Scores"
    data = Reference(ws, min_col=3, min_row=1, max_row=len(students) + 1)
    categories = Reference(ws, min_col=2, min_row=2, max_row=len(students) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 15
    ws.add_chart(chart, "F5")

    # Adding a new sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws["A1"] = "Total Students"
    summary_ws["B1"] = len(students)
    summary_ws["A2"] = "Average Score"
    summary_ws["B2"] = "=AVERAGE('Demo Sheet'!C2:C6)"
    summary_ws["A3"] = "Highest Score"
    summary_ws["B3"] = "=MAX('Demo Sheet'!C2:C6)"

    # Saving the workbook
    wb.save("openpyxl_demo.xlsx")

    # Reading an existing workbook
    wb2 = load_workbook("openpyxl_demo.xlsx")
    ws2 = wb2.active
    print("Contents of the active sheet:")
    for row in ws2.iter_rows(values_only=True):
        print(row)

    # Adding a pie chart on the summary sheet
    summary_ws = wb2["Summary"]
    chart_pie = PieChart()
    pie_data = Reference(summary_ws, min_col=2, min_row=1, max_row=3)
    chart_pie.add_data(pie_data, titles_from_data=False)
    chart_pie.title = "Summary Data"
    summary_ws.add_chart(chart_pie, "D5")

    # Save the updated workbook
    wb2.save("openpyxl_demo_updated.xlsx")


if __name__ == "__main__":
    demonstrate()
